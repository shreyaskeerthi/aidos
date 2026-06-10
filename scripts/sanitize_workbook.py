#!/usr/bin/env python3
"""Create a share-safe fake clone of an infrastructure Excel workbook.

This script preserves workbook structure and formatting while replacing
sensitive-looking text with realistic fake values.
"""

from __future__ import annotations

import argparse
import json
import random
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EMAIL_RE = re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b")
URL_RE = re.compile(r"\bhttps?://[^\s)]+")
IP_RE = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")
CIDR_RE = re.compile(r"\b((?:\d{1,3}\.){3}\d{1,3})/(\d{1,2})\b")
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{7,}\d)")
SERIAL_RE = re.compile(r"\b[A-Z0-9]{8,}\b")
DEVICE_RE = re.compile(r"\b(?:rack|row|pod|leaf|spine|tor|sw|srv|server|node|dc)-?[a-z0-9-]*\b", re.IGNORECASE)


ENTITY_REPLACEMENTS = {
    "openai": "Apex Dynamics",
    "sanjose": "Rivergate",
    "san jose": "Rivergate",
    "idcs": "MDS",
    "client": "Customer",
    "customer": "Customer",
    "younes": "Evan Cole",
}


@dataclass
class FakeContext:
    seed: int
    rng: random.Random = field(init=False)
    email_count: int = 0
    phone_count: int = 0
    ip_counter: int = 1
    ip_map: dict[str, str] = field(default_factory=dict)
    company_used: str = "Apex Dynamics"
    label_counter: int = 0
    explicit_replacements: dict[str, str] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.rng = random.Random(self.seed)

    def fake_email(self) -> str:
        self.email_count += 1
        return f"user{self.email_count:03d}@example.com"

    def fake_phone(self) -> str:
        self.phone_count += 1
        return f"+1-555-010-{self.phone_count:04d}"[-14:]

    def fake_ip(self, original: str) -> str:
        if original in self.ip_map:
            return self.ip_map[original]
        third_octet = (self.ip_counter // 254) % 254
        fourth_octet = (self.ip_counter % 254) + 1
        fake = f"10.200.{third_octet}.{fourth_octet}"
        self.ip_map[original] = fake
        self.ip_counter += 1
        return fake

    def fake_label(self, prefix: str = "asset") -> str:
        self.label_counter += 1
        return f"{prefix}-{self.label_counter:04d}"


def _replace_entities(text: str, company_name: str) -> str:
    out = text
    mapping = {**ENTITY_REPLACEMENTS, "apex dynamics": company_name.lower()}
    for source, target in mapping.items():
        out = re.sub(source, target, out, flags=re.IGNORECASE)
    return out


def _apply_explicit_replacements(text: str, replacements: dict[str, str]) -> str:
    updated = text
    for source, target in replacements.items():
        updated = re.sub(re.escape(source), target, updated, flags=re.IGNORECASE)
    return updated


def _sanitize_text(value: str, ctx: FakeContext) -> str:
    if value.startswith("="):
        return value

    text = value

    text = EMAIL_RE.sub(lambda _m: ctx.fake_email(), text)
    text = URL_RE.sub("https://example.com", text)
    text = PHONE_RE.sub(lambda _m: ctx.fake_phone(), text)

    def _replace_cidr(match: re.Match[str]) -> str:
        ip = match.group(1)
        mask = match.group(2)
        return f"{ctx.fake_ip(ip)}/{mask}"

    text = CIDR_RE.sub(_replace_cidr, text)
    text = IP_RE.sub(lambda m: ctx.fake_ip(m.group(0)), text)

    # Replace serial-like IDs while keeping length/prefix feel.
    def _replace_serial(match: re.Match[str]) -> str:
        src = match.group(0)
        if src.isdigit():
            return src
        return "SN" + "".join(ctx.rng.choice("0123456789ABCDEF") for _ in range(max(6, len(src) - 2)))

    text = SERIAL_RE.sub(_replace_serial, text)

    text = _replace_entities(text, ctx.company_used)
    text = _apply_explicit_replacements(text, ctx.explicit_replacements)

    # Common direct substitutions for infra sheets.
    text = re.sub(r"\bNorthline Logistics\b", ctx.company_used, text, flags=re.IGNORECASE)
    text = re.sub(r"\bnorthline\b", "atlas", text, flags=re.IGNORECASE)

    return text


def _sanitize_text_aggressive(value: str, ctx: FakeContext) -> str:
    """Aggressive anonymization: preserve structure but replace most identifiers/text."""
    if value.startswith("="):
        return value

    text = _sanitize_text(value, ctx)

    def _replace_device(match: re.Match[str]) -> str:
        src = match.group(0)
        if src.lower().startswith("rack"):
            return ctx.fake_label("rack")
        if src.lower().startswith("row"):
            return ctx.fake_label("row")
        if src.lower().startswith("pod"):
            return ctx.fake_label("pod")
        if src.lower().startswith(("leaf", "spine", "tor", "sw")):
            return ctx.fake_label("switch")
        if src.lower().startswith(("srv", "server", "node", "dc")):
            return ctx.fake_label("node")
        return ctx.fake_label("asset")

    text = DEVICE_RE.sub(_replace_device, text)

    # Replace long human-readable strings with deterministic placeholders.
    stripped = text.strip()
    if len(stripped) > 14 and any(ch.isalpha() for ch in stripped):
        # Keep punctuation footprint if possible.
        if ":" in stripped:
            left, _, right = stripped.partition(":")
            text = f"{ctx.fake_label('section')}: {ctx.fake_label('value')}"
        elif "-" in stripped and not re.search(r"\d", stripped):
            text = ctx.fake_label("label")
        elif " " in stripped:
            words = stripped.split()
            text = " ".join(ctx.fake_label("w") for _ in words)
        else:
            text = ctx.fake_label("txt")

    return text


def _sanitize_cell_value(value: Any, ctx: FakeContext, *, aggressive: bool = False) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        if aggressive:
            return _sanitize_text_aggressive(value, ctx)
        return _sanitize_text(value, ctx)
    return value


def _sanitize_workbook_metadata(wb: Any, ctx: FakeContext, aggressive: bool) -> None:
    props = wb.properties
    props.creator = "sanitized"
    props.lastModifiedBy = "sanitized"
    props.title = f"{ctx.company_used} Infrastructure Workbook"
    props.subject = "Sanitized Infrastructure Data"
    props.description = "Synthetic workbook generated for sharing"
    props.keywords = "sanitized, synthetic, fake"
    props.category = "Sanitized"
    if aggressive:
        props.company = ctx.company_used


def _sanitize_sheet_titles(wb: Any, ctx: FakeContext, aggressive: bool) -> None:
    if not aggressive:
        return
    for idx, ws in enumerate(wb.worksheets, start=1):
        ws.title = f"Sheet-{idx:02d}"


def sanitize_workbook(
    input_path: Path,
    output_path: Path,
    seed: int,
    *,
    aggressive: bool = False,
    company_name: str = "Apex Dynamics",
    explicit_replacements: dict[str, str] | None = None,
) -> dict[str, Any]:
    wb = load_workbook(input_path)
    ctx = FakeContext(
        seed=seed,
        company_used=company_name,
        explicit_replacements=explicit_replacements or {},
    )

    _sanitize_workbook_metadata(wb, ctx, aggressive)
    _sanitize_sheet_titles(wb, ctx, aggressive)

    changed_cells = 0
    total_cells = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                total_cells += 1
                original = cell.value
                sanitized = _sanitize_cell_value(original, ctx, aggressive=aggressive)
                if sanitized != original:
                    cell.value = sanitized
                    changed_cells += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "input": str(input_path),
        "output": str(output_path),
        "seed": seed,
        "total_cells": total_cells,
        "changed_cells": changed_cells,
        "mapped_ips": len(ctx.ip_map),
        "company_name": ctx.company_used,
        "aggressive": aggressive,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Sanitize an infrastructure workbook into fake data.")
    parser.add_argument("input", type=Path, help="Path to real workbook (.xlsx/.xlsm/.xls)")
    parser.add_argument("output", type=Path, nargs="?", help="Output fake workbook path")
    parser.add_argument("--seed", type=int, default=42, help="Random seed for deterministic output")
    parser.add_argument(
        "--company-name",
        default="Apex Dynamics",
        help="Realistic fake company name to use throughout the workbook.",
    )
    parser.add_argument(
        "--replace",
        action="append",
        default=[],
        help="Explicit text replacement in the form source=target. Can be repeated.",
    )
    parser.add_argument(
        "--aggressive",
        action="store_true",
        help="Strong anonymization mode (sheet names, metadata, and most free text).",
    )
    parser.add_argument(
        "--report",
        type=Path,
        default=None,
        help="Optional path to write a JSON report",
    )
    args = parser.parse_args()

    input_path = args.input
    if not input_path.exists():
        raise SystemExit(f"Input workbook not found: {input_path}")

    output_path = args.output or input_path.with_name(f"{input_path.stem}.fake{input_path.suffix}")
    replacements: dict[str, str] = {}
    for item in args.replace:
        if "=" not in item:
            raise SystemExit(f"Invalid --replace value: {item}. Expected source=target")
        source, target = item.split("=", 1)
        replacements[source.strip()] = target.strip()

    report = sanitize_workbook(
        input_path,
        output_path,
        args.seed,
        aggressive=args.aggressive,
        company_name=args.company_name,
        explicit_replacements=replacements,
    )

    print(json.dumps(report, indent=2))

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
